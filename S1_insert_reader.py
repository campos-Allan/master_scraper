"""extract info from excel and cross-reference with other excel file
"""
from openpyxl import Workbook, load_workbook
import pandas as pd
import os


def S1_excel_reader(file):
    """read excel files
    """
    PATH = os.getcwd()+'\\files\\'
    df_s1 = pd.read_excel(PATH+file)
    if 'ABERTURA' in file.upper():
        dia = df_s1[df_s1.columns[0]][2].split(' ')[-1]
        df_s1 = df_s1[[df_s1.columns[4], df_s1.columns[6],
                       df_s1.columns[-1]]][1:].fillna(0)
        change_b, change_c = df_s1.iloc[1], df_s1.iloc[2]
        df_s1.iloc[1] = change_c
        df_s1.iloc[2] = change_b
        df_s1 = df_s1.rename(index={1: 'X1', 2: 'Y2', 3: 'Y1'}, columns={
            df_s1.columns[0]: 'descarga_', df_s1.columns[1]: 'mercado_',
            df_s1.columns[2]: 'saldo_'})
        return df_s1, dia
    elif 'DESCARGA' in file.upper():
        df_s1 = df_s1.fillna(0)
        dia = df_s1.iloc[-1]['d'].strftime('%d/%m/%Y')
        df_s1 = df_s1[df_s1['ORIGEM'] != 0]
        df_s1 = df_s1[[df_s1.columns[6], df_s1.columns[10], df_s1.columns[11]]]
        dic_descarga = {}
        for i in range(0, len(df_s1)):
            dic_descarga[df_s1.iloc[i]['ID'] +
                         '\t'+str(int(df_s1.iloc[i]['Quantidade Carregada']))] = df_s1.iloc[i]['Quantidade Descarregada']
        workbook = Workbook()
        workbook = load_workbook(filename='molde.xlsx')
        sheet = workbook['registro']
        ind = 2
        while sheet[f'A{ind}'].value is not None:
            sheet_value = str(sheet[f'E{ind}'].value)+'\t'+str(
                sheet[f'G{ind}'].value)
            if sheet_value in dic_descarga:
                sheet[f'I{ind}'] = dia
                sheet[f'K{ind}'] = int(dic_descarga[sheet_value])
            ind += 1
        workbook.save(filename="molde.xlsx")
        workbook.close()
