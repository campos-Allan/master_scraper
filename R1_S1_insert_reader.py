"""extract info from excel and cross-reference with other excel file
"""
from openpyxl import Workbook, load_workbook
import pandas as pd
import os


def excel_reader(file, registro):
    """read excel files
    """
    # pylint: disable=used-before-assignment
    PATH = os.getcwd()+'\\files\\'
    workbook = Workbook()
    workbook = load_workbook(filename=PATH+file)
    if 'BASIS' in file.upper():
        ind = 1
        sheet = workbook['Planilha1']
        while sheet[f'A{ind}'].value is not None:
            ind += 1
        for j in range(1, ind):
            operador = sheet[f'C{j}'].value
            if (operador == 'R1') or \
                    (operador == 'T2'):
                placa = sheet[f'D{j}'].value
                produto = sheet[f'A{j}'].value
                if produto == "Y1":
                    produto = "Y1"
                elif produto == "Y2":
                    produto = "Y2"
                else:
                    produto = "X1"
                vol_carga = sheet[f'L{j}'].value
                dia = sheet[f'K{j}'].value.date().strftime('%d/%m/%Y')
                nome = sheet[f'G{j}'].value
                if 'R1' in operador:
                    orgao = 'R1'
                    registro.loc[placa] = [dia, 'R1', 'R2', 'MODAL A',
                                           placa, produto, vol_carga, '', '', '', '', nome]
                elif 'T2' in operador:
                    orgao = 'T2'
                    registro.loc[placa] = [dia, 'T2', 'S1', 'MODAL A',
                                           placa, produto, vol_carga, '', '', '', '', nome]
        sheet = workbook['Planilha2']
        old_total = {sheet['A4'].value: str(sheet['B4'].value),
                     sheet['A5'].value: str(sheet['B5'].value),
                     sheet['A6'].value: str(sheet['B6'].value)}
        total = {}
        if orgao == 'R1':
            for i, j in old_total.items():
                if i is None:
                    pass
                elif 'Total' in i:
                    pass
                else:
                    if i == "Y1":
                        i = "saida_Y1"
                    elif i == "Y2":
                        i = "saida_Y2"
                    else:
                        i = "saida_X1"
                    total[i] = [j]
        elif orgao == 'T2':
            for i, j in old_total.items():
                if i is None:
                    pass
                elif 'Total' in i:
                    pass
                else:
                    if i == "Y1":
                        i = "envio_T2_Y1"
                    elif i == "Y2":
                        i = "envio_T2_Y2"
                    else:
                        i = "envio_T2_X1"
                    total[i] = [j]
        total['dia'] = dia
        return total, orgao
    if 'SEN' in file.upper():
        PATH = os.getcwd()+'\\files\\'
        df_dt = pd.read_excel(PATH+file)
        alcunha = 'X1'
        indice = 0
        dic_descarga = {}
        for i in df_dt['Resumo']:
            if i == 'X1':
                alcunha = 'X1'
            elif i == 'Y2':
                alcunha = 'Y2'
            elif i == 'Y1':
                alcunha = 'Y1'
            else:
                df_dt.loc[indice, 'Resumo'] = alcunha
            indice += 1
        descarga = df_dt[df_dt['Unnamed: 4'] == 'DESCARREGAMENTO']

        descarga['Unnamed: 1'] = descarga['Unnamed: 1'].apply(
            lambda x: x.strftime('%d/%m/%Y'))
        descarga['Resumo'] = descarga['Resumo'].replace('X1', 'X1').replace(
            'Y2', 'Y2').replace('Y1', 'Y1')
        descarga['Unnamed: 2'] = descarga['Unnamed: 2'].map(
            lambda x: x.replace('-', ''))
        descarga['Unnamed: 16'] = descarga['Unnamed: 16'].apply(
            str).map(lambda x: x.replace('.', ''))
        descarga['Unnamed: 18'] = descarga['Unnamed: 18'].apply(
            str).map(lambda x: x.replace('.', ''))
        metade_1 = descarga[['Resumo', 'Unnamed: 16', 'Unnamed: 3']]
        metade_2 = descarga[['Unnamed: 1',
                             'Unnamed: 2', 'Unnamed: 18']]
        metade_1 = metade_1.stack().groupby(level=0).apply('\t'.join).to_dict()
        metade_2 = metade_2.stack().groupby(level=0).apply('\t'.join).to_dict()
        for i, j in metade_1.items():
            dic_descarga[j] = metade_2[i]
        workbook = Workbook()
        workbook = load_workbook(filename='molde.xlsx')
        sheet = workbook['registro']
        ind = 2
        used = {}
        while sheet[f'A{ind}'].value is not None:
            sheet_value = str(sheet[f'F{ind}'].value)+'\t'+str(
                sheet[f'G{ind}'].value)+'\t'+str(
                sheet[f'L{ind}'].value).split(' ')[0]
            if sheet_value in dic_descarga:  # pylint: disable=used-before-assignment
                used[sheet_value] = 1
                sheet[f'I{ind}'] = dic_descarga[sheet_value].split('\t')[0]
                sheet[f'J{ind}'] = dic_descarga[sheet_value].split('\t')[1]
                sheet[f'K{ind}'] = dic_descarga[sheet_value].split('\t')[2]
            ind += 1
        workbook.save(filename="molde.xlsx")
        workbook.close()
        nao_usados = dic_descarga.copy()
        for i, _ in used.items():
            del nao_usados[i]
        print('Não achados')
        for i, j in nao_usados.items():
            print(i+'\t'+j)
        return df_dt
