"""pdf scraping and excel formatting
"""
# CRIAR MÉTODO PRA ENUMERAR DESCARREGAMENTOS QUE NÃO FORAM COLOCADOS NA PLANILHA
import time
import datetime
from datetime import timedelta
import os
from os import listdir
from os.path import isfile, join
import warnings
from tkinter import messagebox
import pyautogui
from tabula import read_pdf
from pypdf import PdfReader
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
warnings.filterwarnings("ignore")


def sap(cod: str) -> None:
    """bot to get data from a software

    Args:
        cod (str): sector code, should be 1120 or 1109
    """
    now = datetime.datetime.now()
    ano_sexta = 0
    if now.today().weekday() == 0:
        sexta = now - timedelta(days=3)
        segunda = 'yes'
        ano_sexta = str(sexta.year)
        dia_sexta = str(sexta.day)
        mes_sexta = str(sexta.month)
        if len(mes_sexta) == 1:
            mes_sexta = '0'+mes_sexta

        if len(dia_sexta) == 1:
            dia_sexta = '0'+dia_sexta

    else:
        segunda = 'no'
    ontem = now - timedelta(days=1)
    ano = str(ontem.year)
    dia = str(ontem.day)
    mes = str(ontem.month)

    if len(mes) == 1:
        mes = '0'+mes

    if len(dia) == 1:
        dia = '0'+dia
    pyautogui.click(730, 1050)
    time.sleep(7)
    pyautogui.click(50, 135)
    pyautogui.click(50, 135)
    messagebox.showinfo(
        "Continuar", "Se tiver aparecido um aviso, feche e aperte em ok")
    time.sleep(3)
    pyautogui.click(270, 80)
    pyautogui.click(270, 105)
    pyautogui.click(270, 80)
    pyautogui.click(50, 80)
    pyautogui.click(50, 80)
    time.sleep(2)
    pyautogui.typewrite(cod)
    pyautogui.click(700, 700)
    if segunda == 'yes':
        pyautogui.click(500, 300)
        time.sleep(1)
        pyautogui.typewrite(f'{dia_sexta}.{mes_sexta}.{ano_sexta}')
        pyautogui.click(800, 300)
        time.sleep(1)
        pyautogui.typewrite(f'{dia}.{mes}.{ano}')
    else:
        pyautogui.click(500, 300)
        time.sleep(1)
        pyautogui.typewrite(f'{dia}.{mes}.{ano}')
        pyautogui.click(800, 300)
        time.sleep(1)
        pyautogui.typewrite(f'{dia}.{mes}.{ano}')

    pyautogui.click(630, 480)
    pyautogui.press('backspace', presses=12)
    pyautogui.click(37, 185)
    time.sleep(9)
    pyautogui.click(150, 185)
    time.sleep(1)
    pyautogui.click(60, 35)

    pyautogui.click(100, 130)
    time.sleep(1)
    pyautogui.click(500, 170)
    time.sleep(1)
    pyautogui.click(400, 550)
    time.sleep(2)
    pyautogui.moveTo(400, 755, duration=1)
    pyautogui.moveTo(405, 755, duration=1)
    time.sleep(1)
    pyautogui.click(400, 755)
    time.sleep(2)
    pyautogui.click(680, 720)
    time.sleep(3)
    pyautogui.click(900, 875)
    time.sleep(4)
    pyautogui.click(310, 585)
    time.sleep(3)
    pyautogui.click(760, 685)
    time.sleep(3)
    pyautogui.click(535, 505)
    time.sleep(5)
    pyautogui.getWindowsWithTitle("Planilha em Basis (1)")[0].maximize()
    time.sleep(1)
    if cod == '1120':
        pyautogui.click(650, 105)
        time.sleep(1)
        pyautogui.click(120, 220)
        time.sleep(7)
        pyautogui.click(1300, 640)
        time.sleep(10)
        pyautogui.click(230, 230)
        time.sleep(5)
        pyautogui.click(350, 150)
        time.sleep(2)
        pyautogui.click(1450, 730)
        time.sleep(2)
        pyautogui.moveTo(1370, 530, duration=1)
        pyautogui.click(1370, 530)
        time.sleep(2)
        pyautogui.moveTo(1370, 600)
        pyautogui.mouseDown(button='right')
        pyautogui.mouseUp(button='right')
        time.sleep(1)
        pyautogui.click(1380, 610)
        time.sleep(1)
        pyautogui.scroll(-10000)
        time.sleep(1)
        pyautogui.click()
        time.sleep(1)
        messagebox.showinfo(
            "Sucesso", "Coloque Rio Verde no filtro, salve o arquivo na mesma pasta desse programa")
    else:
        messagebox.showinfo(
            "Sucesso", "Coloque SINOP no filtro, salve o arquivo na mesma pasta desse programa")


def excel_write(dic_descarga: dict, action: str) -> None:
    """write data extracted from pdf and excel files in "registro.xlxs" 

    Args:
        dic_descarga (dict): data extracted
        action (str): type of info to write, should be 'Movimento' or 'Descarga'
    """
    workbook = Workbook()
    workbook = load_workbook(filename='registro.xlsx')
    ind = 1
    for i in workbook.sheetnames:
        sheet = workbook[i]
        if i == 'Log':
            while sheet[f'A{ind}'].value is not None:
                ind += 1
            if action == 'Movimentação':
                for j in range(1, len(dic_descarga.split('\n'))):
                    sheet[f'A{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[0]
                    sheet[f'B{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[1]
                    sheet[f'C{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[2]
                    sheet[f'D{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[3]
                    sheet[f'E{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[4]
                    sheet[f'F{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[5]
                    sheet[f'G{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[6]
                    sheet[f'L{ind+j-1}'] = dic_descarga.split(
                        '\n')[j-1].split('\t')[-1]
            else:
                planilha_valor = ''
                for j in range(1, ind):
                    if len(str(sheet[f'G{j}'].value)) == 6:
                        planilha_valor = str(sheet[f'F{j}'].value)+'\t'+str(
                            sheet[f'G{j}'].value)[:3]+'.'+str(
                            sheet[f'G{j}'].value)[3:]+'\t'+str(sheet[f'L{j}'].value).split(' ')[0]
                    elif len(str(sheet[f'G{j}'].value)) == 5:
                        planilha_valor = str(sheet[f'F{j}'].value)+'\t'+str(
                            sheet[f'G{j}'].value)[:2]+'.'+str(
                            sheet[f'G{j}'].value)[2:]+'\t'+str(sheet[f'L{j}'].value).split(' ')[0]
                    if planilha_valor in dic_descarga:
                        sheet[f'J{j}'] = dic_descarga[planilha_valor].split('\t')[
                            0]
                        sheet[f'I{j}'] = dic_descarga[planilha_valor].split('\t')[
                            1]
                        sheet[f'K{j}'] = dic_descarga[planilha_valor].split('\t')[
                            2]
    workbook.save(filename="registro.xlsx")
    messagebox.showinfo("Sucesso", "Informação inserida com sucesso")


def pdf_reader(operador: str) -> None:
    """read pdf files

    Args:
        operador (str): which operation should be read
    """
    mypath = os.getcwd()
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    arq_ferroviario = [i for i in onlyfiles if 'FERROVIÁRIO' in i.upper()]
    for i in arq_ferroviario:
        onlyfiles.remove(i)
        onlyfiles.append(i)
    realizado = pd.DataFrame(columns=('Dia_carga', 'Origem', 'Destino', 'Modal', 'ID_carga',
                             'Combustível', 'Vol_carga', 'ETA', 'Dia_descarga', 'ID_descarga',
                                      'Vol_descarga', 'Nome'))
    dic_descarga = {}
    modal = ''
    estoque = ''
    df_teciap = pd.DataFrame()
    linhas_descarga = ''
    mov_tct = ''
    carga = ''
    for i in onlyfiles:
        nome = i
        i = i.upper()
        if (operador == "RONDO") & ('ESTOQUE BR' in i):
            # pylint: disable=unsubscriptable-object
            estoque = pd.DataFrame(read_pdf(nome, stream=True)[0])
            estoque.columns = estoque.iloc[0]
            estoque = estoque.drop(0)
            estoque.loc[:, 'Entradas'] = estoque[estoque.columns[0]].str.split(
                ' ', expand=True)[2]
        elif (operador == "RONDO") & ('SALDO' in i):
            df_teciap = pd.DataFrame(read_pdf(nome, stream=True)[0])
            df_teciap = df_teciap[[df_teciap.columns[6],
                                   df_teciap.columns[9], df_teciap.columns[-1]]]
            df_teciap[df_teciap.columns[0]] = df_teciap[df_teciap.columns[0]].replace(
                '0 0', '0').astype(float)
            df_teciap[df_teciap.columns[1]] = df_teciap[df_teciap.columns[1]].replace(
                '0 0', '0').astype(float)
            for i in range(0, len(df_teciap[df_teciap.columns[2]])):
                if len(df_teciap[df_teciap.columns[2]][i]) > 7:
                    df_teciap[df_teciap.columns[2]][i] = df_teciap[df_teciap.columns[2]][i].replace(
                        '.', '', 1)
            df_teciap[df_teciap.columns[2]
                      ] = df_teciap[df_teciap.columns[2]].astype(float)
            b, c = df_teciap.iloc[1], df_teciap.iloc[2]
            df_teciap.iloc[1] = c
            df_teciap.iloc[2] = b
        elif (operador == "RONDO") & ('DESCARGA' in i):
            arquivo = read_pdf(nome, stream=True)
            novo_vol_carga = 0
            novo_vol_descarga = 0
            chave = ''
            linhas_descarga = []
            for i in range(1, len(arquivo)):
                descarga_teciap = pd.DataFrame(arquivo[i])
                if 'S10' in ' '.join(descarga_teciap.columns):
                    chave = 'DSL S10'
                elif 'S500' in ' '.join(descarga_teciap.columns):
                    chave = 'DSL S500'
                elif 'GASOLINA' in ' '.join(descarga_teciap.columns):
                    chave = 'GASOA'
                vol_carga = descarga_teciap[descarga_teciap.columns[-3]][:-1]
                vol_descarga = descarga_teciap[descarga_teciap.columns[-2]][:-1]
                placa = descarga_teciap[descarga_teciap.columns[-4]
                                        ][:-1].str.replace('-', '')
                dia = descarga_teciap[descarga_teciap.columns[0]][:-1]
                motor = descarga_teciap[descarga_teciap.columns[3]][:-1]
                if 'RUMO' in motor[0]:
                    vol_carga = descarga_teciap[descarga_teciap.columns[-4]][:-1]
                    vol_descarga = descarga_teciap[descarga_teciap.columns[-3]][:-1]
                    placa = descarga_teciap[descarga_teciap.columns[-5]
                                            ][:-1].str.replace('-', '')
                for i in range(0, len(descarga_teciap[:-1])):
                    if 'RUMO' in motor[i]:
                        motor[i] = 'VAGAO-INDEFINIDO'
                    if len(str(vol_carga[i])) < 6:
                        novo_vol_carga = str(
                            vol_carga[i])+'0'*(6-len(str(vol_carga[i])))
                        dic_descarga[chave+'\t'+novo_vol_carga+'\t'+motor[i].split(
                            ' ')[0]] = placa[i]+'\t'+dia[i]+'\t'+str(vol_descarga[i])
                        linhas_descarga.append(chave+'\t'+novo_vol_carga+'\t'+motor[i].split(' ')[
                            0]+'\t'+dia[i]+'\t'+placa[i]+'\t'+str(vol_descarga[i]))
                    if len(str(vol_descarga[i])) < 6:
                        novo_vol_descarga = str(
                            vol_descarga[i])+'0'*(6-len(str(vol_descarga[i])))
                        dic_descarga[chave+'\t'+str(vol_carga[i])+'\t'+motor[i].split(
                            ' ')[0]] = placa[i]+'\t'+dia[i]+'\t'+novo_vol_descarga
                        linhas_descarga.append(chave+'\t'+str(vol_carga[i])+'\t'+motor[i].split(
                            ' ')[0]+'\t'+dia[i]+'\t'+placa[i]+'\t'+novo_vol_descarga)
                    if novo_vol_carga == 0 and novo_vol_descarga == 0:
                        dic_descarga[chave+'\t'+str(vol_carga[i])+'\t'+motor[i].split(
                            ' ')[0]] = placa[i]+'\t'+dia[i]+'\t'+str(vol_descarga[i])
                        linhas_descarga.append(chave+'\t'+str(vol_carga[i])+'\t'+motor[i].split(
                            ' ')[0]+'\t'+dia[i]+'\t'+placa[i]+'\t'+str(vol_descarga[i]))
                    novo_vol_carga = 0
                    novo_vol_descarga = 0
            op_saldo = 'N'
            dia = []
            for chave_ferroviario, _ in dic_descarga.items():
                if 'VAGAO-INDEFINIDO' in chave_ferroviario:
                    op_saldo = 'Y'
                    dia.append(dic_descarga[chave_ferroviario].split('\t')[1])
            if op_saldo == 'Y':
                modal = {'fe_GASOA': 0, 'fe_DSL500': 0, 'fe_DSL10': 0,
                         'ro_GASOA': 0, 'ro_DSL500': 0, 'ro_DSL10': 0}
                for k, l in dic_descarga.items():
                    if l.split('\t')[1] == dia[-1]:
                        if k.split('\t')[0] == 'GASOA':
                            if 'VAGAO' in k.split('\t')[2]:
                                modal['fe_GASOA'] = modal['fe_GASOA'] + \
                                    float(l.split('\t')[2])
                            else:
                                modal['ro_GASOA'] = modal['ro_GASOA'] + \
                                    float(l.split('\t')[2])
                        elif k.split('\t')[0] == 'DSL S500':
                            if 'VAGAO' in k.split('\t')[2]:
                                modal['fe_DSL500'] = modal['fe_DSL500'] + \
                                    float(l.split('\t')[2])
                            else:
                                modal['ro_DSL500'] = modal['ro_DSL500'] + \
                                    float(l.split('\t')[2])
                        elif k.split('\t')[0] == 'DSL S10':
                            if 'VAGAO' in k.split('\t')[2]:
                                modal['fe_DSL10'] = modal['fe_DSL10'] + \
                                    float(l.split('\t')[2])
                            else:
                                modal['ro_DSL10'] = modal['ro_DSL10'] + \
                                    float(l.split('\t')[2])
            else:
                modal = ''
        elif (operador == "RONDO") & ('MOV BR' in i):
            mov_tct = ''
            vagao = ''
            reader = PdfReader(nome)
            number_of_pages = len(reader.pages)
            for j in range(0, number_of_pages):
                page = reader.pages[j]
                if 'Veículo' in page.extract_text():
                    text = page.extract_text().split(
                        'Motorista')[-1].split('CONTROLE E EXPEDIÇÃO DE COMBUSTÍVEIS')[0]
                    text = text[text.find('/')-2:]
                    dia = text[:10]
                    text = text.split(dia)
                    for k in range(0, len(text)-1):
                        dados = text[k+1].split('\n')
                        placa = dados[2]
                        if placa == 'VAGAO':
                            vagao = dados[1]
                        nome = dados[3].strip()
                        if len(placa) != 7:
                            placa = dados[2]+dados[3]
                            nome = dados[4]
                        vol = dados[-2]
                        if vol.find(',') != -1:
                            vol = dados[-3]
                            if vol.find(',') != -1:
                                vol = dados[-4]
                        if j == 0 or j == 3:
                            if placa == 'VAGAOVAGAO-INDEFINIDO':
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'ferroviário' + '\t'+vagao+'\t'+'GASOA'+'\t'+vol+'\t'+'\t' + \
                                    '\t'+'\t'+'\t'+'VAGAO-INDEFINIDO'+'\n'
                            else:
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'rodoviário' + '\t'+placa+'\t'+'GASOA'+'\t'+vol+'\t'+'\t' \
                                    + '\t'+'\t'+'\t'+nome+'\n'
                        if j == 1 or j == 4:
                            if placa == 'VAGAOVAGAO-INDEFINIDO':
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'ferroviário'+'\t' + vagao+'\t'+'DSL S10'+'\t'+vol+'\t'+'\t' + \
                                    '\t'+'\t'+'\t'+'VAGAO-INDEFINIDO'+'\n'
                            else:
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'rodoviário' + '\t'+placa+'\t'+'DSL S10'+'\t'+vol+'\t'+'\t' + \
                                    '\t' + '\t'+'\t'+nome+'\n'
                        if j == 2 or j == 5:
                            if placa == 'VAGAOVAGAO-INDEFINIDO':
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'ferroviário'+'\t' + vagao+'\t'+'DSL S500'+'\t'+vol+'\t' + \
                                    '\t'+'\t'+'\t'+'\t'+'VAGAO-INDEFINIDO'+'\n'
                            else:
                                mov_tct = mov_tct+dia+'\t'+'Paulínia'+'\t'+'Rondonópolis'+'\t' + \
                                    'rodoviário'+'\t'+placa+'\t'+'DSL S500'+'\t'+vol+'\t' + \
                                    '\t'+'\t'+'\t'+'\t'+nome+'\n'
                        realizado.loc[len(realizado)] = mov_tct.split(
                            '\n')[-2].split('\t')
            realizado["Vol_carga"] = realizado["Vol_carga"].astype(float)
            df = pd.DataFrame(realizado.groupby(['Modal', 'Combustível']).sum()[
                              "Vol_carga"]).round(3).reset_index()
            carga = pd.DataFrame(data=(0), index=(
                'ferroviário', 'rodoviário'), columns=('GASOA', 'DSL S500', 'DSL S10'))
            for i in range(0, len(df)):
                carga[df.iloc[i]['Combustível']][df.iloc[i]['Modal']] = str(
                    df['Vol_carga']).split('\n')[i].split(' ')[-1]
            carga.replace(to_replace=0.000, value='0', inplace=True)
            carga = carga.astype(str)
            realizado = pd.DataFrame(columns=('Dia_carga', 'Origem', 'Destino', 'Modal', 'ID_carga',
                                     'Combustível', 'Vol_carga', 'ETA', 'Dia_descarga',
                                              'ID_descarga', 'Vol_descarga', 'Nome'))
    return (estoque, df_teciap, dic_descarga, linhas_descarga, modal, mov_tct, carga)


def excel_reader(operador: str) -> None:
    """read excel files

    Args:
        operador (str): which operation should be read
    """
    try:
        workbook = Workbook()
        dicionario_totais = {}
        nova_descarga = ''
        if operador == 'SINOP':
            workbook = load_workbook(filename='Planilha em Basis (2).xlsx')
            mov_sinop = ''
        else:
            workbook = load_workbook(filename='Planilha em Basis (1).xlsx')
            mov_dtc = ''
        for i in workbook.sheetnames:
            ind = 1
            sheet = workbook[i]
            if i == 'Planilha2':
                dicionario_totais = {sheet['A4'].value: str(sheet['B4'].value),
                                     sheet['A5'].value: str(sheet['B5'].value),
                                     sheet['A6'].value: str(sheet['B6'].value)}
            elif i == 'Planilha1':
                while sheet[f'A{ind}'].value is not None:
                    ind += 1
                for j in range(1, ind):
                    if (sheet[f'C{j}'].value == 'DINAMICA TERMINAIS RIO VERDE S/A') or \
                            (sheet[f'C{j}'].value == 'SINOP'):
                        placa = sheet[f'D{j}'].value
                        produto = sheet[f'A{j}'].value
                        if produto == "PB.6DH":
                            produto = "DSL S10"
                        elif produto == "PB.658":
                            produto = "DSL S500"
                        else:
                            produto = "GASOA"
                        vol_carga = sheet[f'L{j}'].value
                        dia = sheet[f'K{j}'].value.date().strftime('%d/%m/%Y')
                        motorista = sheet[f'G{j}'].value
                        if operador == 'RV':
                            mov_dtc = mov_dtc + \
                                f'{dia}\tSenador Canedo\tRio Verde\trodoviário\t{
                                    placa}\t{produto}\t{vol_carga}\t\t\t\t\t{motorista}\n'
                        elif operador == 'SINOP':
                            mov_sinop = mov_sinop + \
                                f'{dia}\tRondonópolis\tSinop\trodoviário\t{placa}\t{
                                    produto}\t{vol_carga}\t\t\t\t\t{motorista}\n'
    except FileNotFoundError:
        return 'Sem entradas'
    if operador == 'SINOP':
        return (dicionario_totais, mov_sinop)
    mypath = os.getcwd()
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    dic_descarga = {}
    for i in onlyfiles:
        nome = i
        i = i.upper()
        if 'SENADOR' in i:
            df_dtc = pd.read_excel(nome)
            alcunha = 'GAS A'
            indice = 0
            for i in df_dtc['Resumo']:
                if i == 'GAS A':
                    alcunha = 'GAS A'
                elif i == 'S500 A':
                    alcunha = 'S500 A'
                elif i == 'S10 A':
                    alcunha = 'S10 A'
                else:
                    df_dtc.loc[indice, 'Resumo'] = alcunha
                indice += 1
            descarga = df_dtc[df_dtc['Unnamed: 4'] == 'DESCARREGAMENTO']

            descarga['Unnamed: 1'] = descarga['Unnamed: 1'].apply(
                lambda x: str(x.day)+'/'+str(x.month)+'/'+str(x.year))
            descarga['Resumo'] = descarga['Resumo'].replace('GAS A', 'GASOA').replace(
                'S500 A', 'DSL S500').replace('S10 A', 'DSL S10')
            descarga['Unnamed: 2'] = descarga['Unnamed: 2'].map(
                lambda x: x.replace('-', ''))
            descarga['Unnamed: 16'] = descarga['Unnamed: 16'].apply(str)
            descarga['Unnamed: 18'] = descarga['Unnamed: 18'].apply(str)
            descarga['Unnamed: 16'] = descarga['Unnamed: 16'].map(
                lambda x: x[:2]+'.'+x[2:])
            descarga['Unnamed: 18'] = descarga['Unnamed: 18'].map(
                lambda x: x[:2]+'.'+x[2:])
            nova_descarga = descarga[[
                'Resumo', 'Unnamed: 16', 'Unnamed: 3', 'Unnamed: 1', 'Unnamed: 2',
                'Unnamed: 18']]
            metade_1 = descarga[['Resumo', 'Unnamed: 16', 'Unnamed: 3']]
            metade_2 = descarga[['Unnamed: 1',
                                 'Unnamed: 2', 'Unnamed: 18']]
            metade_1 = metade_1.stack().groupby(level=0).apply('\t'.join).to_dict()
            metade_2 = metade_2.stack().groupby(level=0).apply('\t'.join).to_dict()
            for i, j in metade_1.items():
                dic_descarga[j] = metade_2[i]
    return (dicionario_totais, mov_dtc, df_dtc, nova_descarga, dic_descarga)
